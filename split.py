import pandas as pd
import numpy as np
import os
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

print("Bienvenido al programa de división de archivos Excel")



# Recibir la ruta del archivo Excel original
ruta = input("Introduce la ruta del archivo Excel original: ")


# Recibir la ruta donde se va guardar los archivos Excel
rutaDestino = input("Introduce la ruta donde se va guardar los archivos Excel: ")

# Formatear la ruta segun el sistema operativo
ruta = os.path.normpath(ruta)
rutaDestino = os.path.normpath(rutaDestino)
rutaDestino = os.path.join(rutaDestino, "")

# Recibir el número de filas por DataFrame
filas_por_dataframe = int(input("Introduce el número de filas por DataFrame: "))

# Leer el archivo Excel original
df = pd.read_excel(ruta)

# Recibir que recibira cada archivo
nombre = input("Introduce el nombre que recibira cada archivo: ")

# Calcula la cantidad de DataFrames necesarios
num_dataframes = len(df) // filas_por_dataframe + 1

# Divide el DataFrame en varios DataFrames
dfs = []
for i in range(num_dataframes):
  inicio = i * filas_por_dataframe
  fin = (i + 1) * filas_por_dataframe
  df_dividido = df.iloc[inicio:fin]
  dfs.append(df_dividido)

# Guardar cada DataFrame en un archivo Excel
for i, df_split in enumerate(dfs):
  # Crear un nuevo archivo Excel con openpyxl
  writer = pd.ExcelWriter(f'{rutaDestino}{nombre}_{i+1}.xlsx', engine='openpyxl')
  
  # Convertir el DataFrame a Excel
  df_split.to_excel(writer, index=False)
  
  # Cargar el libro de trabajo
  workbook = writer.book
  
  # Aplicar el estilo a cada columna en cada hoja
  for worksheet in workbook.worksheets:
      for col in worksheet.columns:
          max_length = 0
          column = col[0].column 
          for cell in col:
              try: 
                  if len(str(cell.value)) > max_length:
                      max_length = len(cell.value)
              except:
                  pass
          adjusted_width = (max_length + 2)
          worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width
  
  # Guardar el archivo Excel
  writer._save()

print("Se han guardado todos los archivos")